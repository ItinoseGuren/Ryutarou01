import openpyxl
import requests
from bs4 import BeautifulSoup


url = "https://atinn.jp/area/tokai-tokai"
r = requests.get(url)
soup = BeautifulSoup(r.content, "html.parser")


for prohis_box in soup.find_all(class_="prohis_box"):	#find_all / findは１つ上のものだけ。
	
	titles = prohis_box.find(class_="prohis_boxttl").get_text()
	
	prices = prohis_box.find(class_="current_price price_day2").get_text()

	flames = prohis_box.find(class_="prohis_boxintb2_td2a").get_text()


#ターミナル上に一旦欲しいデータ printで表示させるのが安定。
# ctrl + / で一斉にコメントアウトできる
# ctrl + shift + iで開発者ツールを開ける。

book = openpyxl.load_workbook('Ryutarou01.xlsx')
sheet = book.active

START_ROW_NO = 3
FINISH_ROW_NO = 10

TITLE_COLUMN_NO = 1
PRICE_COLUMN_NO = 2
FLAME_COLUMN_NO = 3

max_row = sheet.max_row

for row_no in range(START_ROW_NO, max_row+1):  #max_row属性はデータの入力されている最大のrow番号。 
	age = sheet.cell(row_no, TITLE_COLUMN_NO).value
	title_write = ''

	if title_write == None:
		break
	else:
		title_write == titles

for row_no in range(START_ROW_NO, max_row+1):   
	age = sheet.cell(row_no, PRICE_COLUMN_NO).value
	price_write = ''

	if price_write == None:
		break
	else:
		price_write == prices


for row_no in range(START_ROW_NO, max_row+1):   
	age = sheet.cell(row_no, FLAME_COLUMN_NO).value
	flame_write = ''

	if flame_write == None:
		break
	else:
		flame_write == flames

book.save('Ryutarou01.xlsx')

book.close

# start .で指定先のフォルダー（エクスプローラー）を開く